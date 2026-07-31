"""Excel 导出器 — 格式化输出（表头加粗、列宽自适应、冻结首行）"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import pandas as pd
from src.export.base import BaseExporter, ExportError


class ExcelExporter(BaseExporter):
    """Excel 格式导出器（.xlsx）"""

    extension = ".xlsx"
    file_filter = "Excel 文件 (*.xlsx)"

    # 样式常量
    HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    CELL_FONT = Font(name="Microsoft YaHei", size=10)
    CELL_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    THIN_BORDER = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    def __init__(self, df: pd.DataFrame, options: dict = None):
        super().__init__(df, options or {})
        self.sheet_name = self.options.get("sheet_name", "Sheet1")
        self.freeze_panes = self.options.get("freeze_panes", True)

    def export(self, path: str) -> str:
        path = self.ensure_extension(path, ".xlsx")
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name[:31]  # Excel 限制 31 字符

            # 写入表头
            for col_idx, col_name in enumerate(self.df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=str(col_name))
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = self.HEADER_ALIGNMENT
                cell.border = self.THIN_BORDER

            # 写入数据
            for row_idx, (_, row) in enumerate(self.df.iterrows(), 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx,
                                   value=self._format_cell(value))
                    cell.font = self.CELL_FONT
                    cell.alignment = self.CELL_ALIGNMENT
                    cell.border = self.THIN_BORDER

            # 列宽自适应（取前 100 行 + 表头的最长值）
            for col_idx in range(1, len(self.df.columns) + 1):
                col_letter = get_column_letter(col_idx)
                max_length = len(str(self.df.columns[col_idx - 1]))
                sample = self.df.iloc[:100, col_idx - 1].dropna().astype(str)
                if len(sample) > 0:
                    max_length = max(max_length, sample.str.len().max())
                # 中文字符约占 2 个字符宽度
                adjusted = min(max_length * 2 + 4, 60)
                ws.column_dimensions[col_letter].width = max(adjusted, 8)

            # 冻结首行
            if self.freeze_panes:
                ws.freeze_panes = "A2"

            # 行高
            ws.row_dimensions[1].height = 28
            for row_idx in range(2, min(len(self.df) + 2, 102)):
                ws.row_dimensions[row_idx].height = 22

            wb.save(path)
            return path

        except Exception as e:
            raise ExportError(f"Excel 导出失败: {e}") from e

    @staticmethod
    def _format_cell(value) -> object:
        """格式化单元格值"""
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return ""
        if isinstance(value, float):
            # 整数显示为整数
            if value == int(value):
                return int(value)
            # 保留 4 位小数
            return round(value, 4)
        return value
